import sys, os, math, datetime, shutil
from timeit import default_timer as timer
import re, unicodedata
from nltk.tag import pos_tag 
from nltk import word_tokenize
from nltk.stem.porter import *
from nltk.stem import RSLPStemmer
from nltk.corpus import stopwords
from pymongo import MongoClient
from openpyxl import load_workbook

from pyspark import SparkContext

from pyspark.ml.feature import HashingTF, IDF
from pyspark.mllib.regression import LabeledPoint
from pyspark.mllib.linalg import SparseVector
from pyspark.mllib.classification import SVMWithSGD, SVMModel


from pyspark.sql import Row, SQLContext

from pyspark.mllib.classification import NaiveBayes, NaiveBayesModel

def createMongoDBConnection(host, port, username, password, db):
	""" create connection with MongoDB
    Args:
        params to connection
    Returns:
        a connection to db
    """
	client = MongoClient(host, port)
	return client[db]

def removeAccents(s):
  s = ''.join((c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn'))
  return re.sub(r'[^\w]', ' ', s)

def findProductsByCategory(categories):
    
    db = createMongoDBConnection(host, port, username, password, database)
    produtos = db.produto_novo
    
    product_list = []
    query_filter = {}
    if categories:
      query_filter = {"categorias" : {"$in" : categories}}
    
    for produto in produtos.find(query_filter):
        keys = produto.keys()
        description = ''
        if 'descricaoLonga' in keys:
            description = removeAccents(description + produto['descricaoLonga'])
        if 'nome' in keys:
            description = removeAccents(description + produto ['nome'])
        id = None
        if '_id' in keys:
            id = str(produto['_id'])

        category = ''
        subcategory = ''

        if 'categorias' in keys:
            categorias_0 = removeAccents(produto['categorias'][0])

            if categorias_0 == "Inicio":            
                category = removeAccents(produto['categorias'][1])
                if(len(produto['categorias']) > 2):
                    subcategory = removeAccents(produto['categorias'][2])
            else:
                category = categorias_0
                if(len(produto['categorias']) > 1):
                    subcategory = removeAccents(produto['categorias'][1])

        if len(description)>0:
            product_list.append((id, description, category, subcategory))
    
    return product_list

def insertTokensAndCategories(tokens, category, categoryAndSubcategory):
    db = createMongoDBConnection(host, port, username, password, database)

    modelCollection = db.model
    modelCollection.remove({'_type':'token'})

    document_mongo =  dict()
    document_mongo['_type'] = 'token'
    document_mongo['_datetime'] = datetime.datetime.utcnow()
    i = 0
    for t in tokens:
        document_mongo[t] = i
        i = i + 1   

    modelCollection.insert_one(document_mongo)

    modelCollection.remove({'_type':'category'})

    document_mongo =  dict()
    document_mongo['_type'] = 'category'
    document_mongo['_datetime'] = datetime.datetime.utcnow()
    i = 0
    for c in category:
        document_mongo[c] = i
        i = i + 1 

    modelCollection.insert_one(document_mongo)

    modelCollection.remove({'_type':'category and subcategory'})
    
    document_mongo =  dict()
    document_mongo['_type'] = 'category and subcategory'
    document_mongo['_datetime'] = datetime.datetime.utcnow()
    i = 0
    for c in categoryAndSubcategory:
        document_mongo[c[0]+","+c[1]] = i
        i = i + 1 

    modelCollection.insert_one(document_mongo)


def main(sc, sqlContext):
    start = timer()

    stpwrds = stopwords.words('english')
    tbl_translate = dict.fromkeys(i for i in xrange(sys.maxunicode) if unicodedata.category(unichr(i)).startswith('S') or unicodedata.category(unichr(i)).startswith('P') or unicodedata.category(unichr(i)).startswith('N'))

    print '---Pegando produtos---'
    start_i = timer()
    productRDD = sc.parallelize(findProductsByCategory([]))
    print '####levou %d segundos' % (timer()-start_i)

    print '---Criando corpus---'
    start_i = timer()
    corpusRDD = (productRDD.map(lambda s: (s[0], word_tokenize(s[1].translate(tbl_translate).lower()), s[2], s[3]))
                           .map(lambda s: (s[0], [PorterStemmer().stem(x) for x in s[1] if x not in stpwrds], s[2], s[3] ))
                           .map(lambda s: (s[0], [x[0] for x in pos_tag(s[1]) if x[1] == 'NN' or x[1] == 'NNP'], s[2], s[3]))
                           .cache())
    print '####levou %d segundos' % (timer()-start_i)

    print '---Pegando e persistindo dados de categoria e tokens---'
    start_i = timer()
    tokens = corpusRDD.flatMap(lambda x: x[1]).distinct().collect()
    numTokens = len(tokens)
    category = productRDD.map(lambda x: x[2]).distinct().collect()
    categoryAndSubcategory = productRDD.map(lambda x: (x[2], x[3])).distinct().collect()
    insertTokensAndCategories(tokens, category, categoryAndSubcategory)
    print '####levou %d segundos' % (timer()-start_i)    

    print '---Calculando TF-IDF dos produtos---'
    start_i = timer()
    wordsData = corpusRDD.map(lambda s: Row(label=s[0], words=s[1], category=s[2], subcategory=s[3]))
    #persistir isso para que ele nao tenha que fazer de novo na predicaoo
    wordsDataDF = sqlContext.createDataFrame(wordsData)   

    #persistindo para a predicao
    wordsDataForPrediction = corpusRDD.map(lambda s: Row(label=s[0], words=s[1], type=s[2]))
    #persistir isso para que ele nao tenha que fazer de novo na predicaoo
    wordsDataForPredictionDF = sqlContext.createDataFrame(wordsDataForPrediction)   

    if os.path.exists("/home/ubuntu/recsys-tcc-ml/parquet/wordsDataDF.parquet"):
        shutil.rmtree("/home/ubuntu/recsys-tcc-ml/parquet/wordsDataDF.parquet")

    wordsDataForPredictionDF.write.parquet("/home/ubuntu/recsys-tcc-ml/parquet/wordsDataDF.parquet") 

    hashingTF = HashingTF(inputCol="words", outputCol="rawFeatures", numFeatures=numTokens)
    idf = IDF(inputCol="rawFeatures", outputCol="features")

    featurizedData = hashingTF.transform(wordsDataDF)
    idfModel = idf.fit(featurizedData)
    rescaledData = idfModel.transform(featurizedData)

    #VSM = rescaledData.map(lambda t: LabeledPoint(categoryAndSubcategory.index((t.category, t.subcategory)), t.features))
    VSM = rescaledData.map(lambda t: LabeledPoint(category.index(t.category), t.features))

    VSMTrain, VSMTest = VSM.randomSplit([8, 2], seed=0L)
    print '####levou %d segundos' % (timer()-start_i)    


    print '--Criando modelo Naive Bayes---'
    start_i = timer()
    model = NaiveBayes.train(VSMTrain)

    if os.path.exists("/home/ubuntu/recsys-tcc-ml/models/naivebayes/modelo_categoria"):
        shutil.rmtree("/home/ubuntu/recsys-tcc-ml/models/naivebayes/modelo_categoria")

    model.save(sc, '/home/ubuntu/recsys-tcc-ml/models/naivebayes/modelo_categoria')
    print '####levou %d segundos' % (timer()-start_i)    

    print '---Testando modelo Naive Bayes---'
    start_i = timer()
    prediction = VSMTest.map(lambda p : (categoryAndSubcategory[int(model.predict(p.features))], categoryAndSubcategory[int(p.label)]))
    acuraccy = float(prediction.filter(lambda (x, v): x[0]==v[0]).count())/float(prediction.count())
    print 'acuracidade de %f' % acuraccy
    print '####levou %d segundos' % (timer()-start_i)    
    
    print '---Pegando os posts---'

    start_i = timer()
    posts = list()
    wb = load_workbook(filename = '/home/ubuntu/recsys-tcc-ml/base_sentimentos.xlsx')
    sheet = wb['Menes']
    for row in sheet.iter_rows(row_offset=1):
        post = list()
        for cell in row:
            if cell.value is None:
                break
            post.append(1 if cell.value == 'Positive' or cell.value == 'Neutral' else 0 if cell.value == 'Negative' else removeAccents(cell.value))

        if len(post) > 0:            
            posts.append(tuple(post))

    print '####levou %d segundos' % (timer()-start_i)

    print '---Criando corpus---'
    start_i = timer()
    postsRDD = sc.parallelize(posts)
    postCorpusRDD = (postsRDD.map(lambda s: (s[1], word_tokenize(s[0].translate(tbl_translate).lower())))
                           .map(lambda s: (s[0], [PorterStemmer().stem(x) for x in s[1] if x not in stpwrds]))
                           .map(lambda s: (s[0], [x[0] for x in pos_tag(s[1]) if x[1] == 'NN' or x[1] == 'NNP']))
                           .cache())

    print '####levou %d segundos' % (timer()-start_i)

    print '---Calculando TF-IDF dos Posts---'
    start_i = timer()
    wordsData = postCorpusRDD.map(lambda s: Row(label=s[0], words=s[1]))
    wordsDataDF = sqlContext.createDataFrame(wordsData)

    hashingTF = HashingTF(inputCol="words", outputCol="rawFeatures", numFeatures=numTokens)
    idf = IDF(inputCol="rawFeatures", outputCol="features")

    featurizedData = hashingTF.transform(wordsDataDF)
    idfModel = idf.fit(featurizedData)
    rescaledData = idfModel.transform(featurizedData)   

    VSM = rescaledData.map(lambda t: LabeledPoint(t.label, t.features))
    VSMTrain, VSMTest = VSM.randomSplit([8, 2], seed=0L)
    print '####levou %d segundos' % (timer()-start_i)   

    print '--Criando modelo SVM---'
    start_i = timer()
    model = SVMWithSGD.train(VSMTrain, iterations=100)
    
    if os.path.exists("/home/ubuntu/recsys-tcc-ml/models/svm"):
        shutil.rmtree("/home/ubuntu/recsys-tcc-ml/models/svm")

    model.save(sc, "/home/ubuntu/recsys-tcc-ml/models/svm")

    print '---Testando modelo SVM---'
    start_i = timer()
    prediction = VSMTest.map(lambda p: (p.label, model.predict(p.features)))
    acuraccy = prediction.filter(lambda (v, p): v != p).count() / float(prediction.count())
    
    print 'acuracidade de %f' % acuraccy

    print '####levou %d segundos' % (timer()-start_i)   

    print 'O processo todo levou %d segundos' % (timer()-start)
    

if __name__ == "__main__":

    host = 'localhost'
    port = 27017
    username = ''
    password = ''
    database = 'tcc-recsys-mongo'

    APP_NAME = 'Recomender System - Treinamento dos Modelos'

    sc = SparkContext(appName=APP_NAME)
    sqlContext = SQLContext(sc)


    main(sc, sqlContext)
    sc.stop()